from openpyxl import load_workbook, Workbook
from correct_client import fix_client_data
from utils import correct_date, create_workbook
import sys
import os

# This is the file that will be corrected
masterFile = sys.argv[1]

# This is the file that will be getting corrected
# data_only flag will only bring in the data and no excel functions
file = load_workbook(filename=masterFile, data_only=True)

client_sheet = file["Clients"]
case_sheet = file["Cases"]
session_sheet = file["Session"]

# This will be a new workbook that we save our corrected values to
workbook = create_workbook()
temp_client_sheet = workbook["Clients"]
temp_case_sheet = workbook["Case"]
temp_session_sheet = workbook["Session"]

# Create a hashtable to keep track of clients
clients = {}
cases = {}
sessions = {}

# Fill our clients hash table with every client
for row in client_sheet.iter_rows(min_row=2, values_only=True, min_col=1):
    client_id = row[1]
    # TODO: IF THE HEADER COLUMNS ARE IN DIFFERENT LOCATIONS, UPDATE THE VALUES FOR THE CLIENT!!!!!
    # This client matches our Template DO NOT CHANGE end up creating multiple
    # clients for each different cms we are importing.
    client = {
        'clientId': client_id,
        'ClientCaseStatus': row[40],
        'ClientProgramEnrollment': row[32],
        'ActiveStaff': row[26],
        'ClientFirstName': row[0],
        'ClientMiddleName': None,
        'ClientLastName': None,
        'DateOfBirth': None,
        'Gender': row[7],
        'Race': row[9],
        'Ethnicity': row[11],
        'VeteranStatus': None,
        'ActiveMilitary': None,
        'FirstTimeHomebuyer': None,
        'HouseholdSize': row[23],
        'CountyAmiIncomeLimit': None,
        'HouseholdIncome': row[17],
        'HouseholdIncomeBand': row[14],
        'IntakeDate': row[28],
        'StreetNumber': None,
        'StreetName': row[2],
        'ApartmentNumber': None,
        'ClientCity': row[3],
        'ClientCounty': row[4],
        'ClientState': row[5],
        'ClientZip': row[6],
        'PrivacyOptOut': None,
        'RuralAreaStatus': row[12],
        'EnglishProficiencyLevel': row[10],
        'BillToHud': None,
        '8a': None,
        '8b': None,
        '8c': None,
        '8d': None,
        '8e': None,
        '8f': None,
        '8g': None,
        '8h': None,
        '8i': None,
        '9a': None,
        '9b': None,
        '9c': None,
        '9d': None,
        '9e': None,
        '9f': None,
        '10a': None,
        '10b': None,
        '10c': None,
        '10d': None,
        '10e': None,
        '10f': None,
        '10g': None,
        '10h': None,
        '10i': None,
        '10j': None,
        '10k': None,
        '10l': None,
        '10m': None,
        'PhoneNumberMobile': None,
        'PhoneNumberWork': None,
        'PhoneNumberHome': None,
        'ImmigrationStatus': None,
        'EmailHome': None,
        'EmailWork': None,
        'MaritalStatus': row[8],
        'Disability': row[22],
        'HouseholdType': row[13],
        'Education': row[18],
        'ReferralSource': row[34],
        'LastContact': None,
        'ActiveReportDateHUD': None,
        'CompletedDate': None,
        'InactiveDate': None
    }

    # Save the client by the ID for easy Access
    clients[client_id] = client

# Add each client to the new spreadsheet
for client in clients:
    client_list = [v for k, v in clients[client].items()]

    # TODO: Fix client data to match our requirements
    # If the address needs split, pass True along with the client list
    # If correcting address use this variable instead
    corrected_client = fix_client_data(client_list, True)
    # If the address does not need split this is the one to use
    # corrected_client = fix_client_data(client_list)

    # Add corrected client to the worksheet
    temp_client_sheet.append(corrected_client)


# ******************************************************************
for row in case_sheet.iter_rows(min_row=2, values_only=True, min_col=1):
    # TODO: IF THE HEADER COLUMNS ARE IN DIFFERENT LOCATIONS, UPDATE THE VALUES FOR THE CLIENT!!!!!
    client_id = row[0]
    case_id = row[5]
    case = {
        'case_type': row[4],
        'client_id': client_id,
        'assigned_counselor': row[1],
        'assigned_coach': None,
        'assigned_loan_officer': None,
        'home_purchase_client_type': None,
        'home_purchase_client_facilitation': None,
        'client_case_status': None,
        'client_disclosure_form_present': None,
        'client_first_name': None,
        'client_middle_name': None,
        'client_last_name': None,
        'date_of_birth': None,
        'credit_score_before': None,
        'credit_score_after': None,
        'intake_date': None,
        'subsidized_housing_assistance': None,
        'primary_employer': None,
        'employer_address': None,
        'secondary_employer': None,
        'secondary_employer_address': None,
        'home_owner_last_three_years': None,
        'real_estate_agent': None,
        'last_contact_date': None,
        'long_term_client_date': None,
        'short_term_client_date': None,
        'near_mortgage_ready_date': None,
        'mortgage_ready_date': None,
        'in_financing_date': None,
        'active_report_date_hud': None,
        'completed_date': None,
        'denied_date': None,
        'inactive_date': None,
        'privacy_opt_out': None,
        'rental_resolution': None,
        'lm_package_status': None,
        'mm_subject_property_present': None,
        'mm_lien_info_present': None,
        'level_one_date': None,
        'level_two_date': None,
        'seeking_shelter_resolution': None,
        'years_at_current_address': None,
        'senior_as_hoh': None,
        'home_owner_resolutions': None,
        'home_purchase_resolution': None
    }
    # Save the client by the ID for easy Access
    cases[case_id] = case

# Add each client to the new spreadsheet
for case in cases:
    case_list = [v for k, v in cases[case].items()]

    # # TODO: Fix client data to match our requirements
    # corrected_case = fix_case_data(case_list)

    # Add corrected client to the worksheet
    temp_case_sheet.append(case_list)


# ******************************************************************
for row in session_sheet.iter_rows(min_row=2, values_only=True, min_col=2):
    session_id = 1
    # TODO: IF THE HEADER COLUMNS ARE IN DIFFERENT LOCATIONS, UPDATE THE VALUES FOR THE CLIENT!!!!!
    session = {
        'session_duration': row[5],
        'counselor_name': row[2],
        'client_notes': None,
        '9_series': row[6],
        '10a': None,
        '10b': None,
        '10c': None,
        '10d': None,
        '10e': None,
        '10f': None,
        '10g': None,
        '10h': None,
        '10i': None,
        '10j': None,
        '10k': None,
        '10l': None,
        '10m': None,
        'session_fee': None,
        'billable_to': None,
        'if_other_who': None,
    }
    # Save the client by the ID for easy Access
    sessions[session_id] = session
    session_id += 1

for session in sessions:
    session_list = [v for k, v in sessions[session].items()]

    # TODO: Fix client data to match our requirements
    # corrected_session = fix_data(session_list)

    # Add corrected client to the worksheet
    temp_session_sheet.append(session_list)

# Save the worksheet when all is complete
outputFileName = os.path.splitext(masterFile)[0] + "_modified.xlsx"
workbook.save(filename=outputFileName)
