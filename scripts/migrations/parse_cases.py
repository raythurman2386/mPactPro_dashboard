from openpyxl import load_workbook
from utils import correct_date

# This is the file that will be corrected
# masterFile = sys.argv[1]

# This is the file that will be getting corrected
# file = load_workbook(filename=masterFile, data_only=True)

# case_sheet = file.active

# This will be a new workbook that we save our corrected values to
workbook = load_workbook(filename='gastonia.xlsx', data_only=True)
new_wb = load_workbook(filename='gastonia_modified.xlsx')
sheet = workbook['Case']
new_sheet = new_wb.create_sheet('Case')

# Create a hashtable to keep track of cases
cases = {}
case_id = 1

template_header = ['CaseType', 'ClientID', 'AssignedCounselor', 'AssignedCoach', 'AssignedLoanOfficer',
                   'HomePurchaseClientType', 'HomePurchaseClientFacilitation', 'ClientCaseStatus',
                   'ClientDisclosureFormPresent', 'ClientFirstName', 'ClientMiddleName', 'ClientLastName',
                   'DateofBirth', 'CreditScoreBefore', 'CreditScoreAfter', 'IntakeDate', 'SubsidizedHousingAssistance',
                   'PrimaryEmployer', 'EmployerAddress', 'SecondaryEmployer', 'SecondaryEmployerAddress',
                   'HomeOwnerLastThreeYears', 'RealEstateAgent', 'LastContactDate', 'LongTermClientDate',
                   'ShortTermClientDate', 'NearMortgageReadyDate', 'MortgageReadyDate', 'InFinancingDate',
                   'ActiveReportDateHUD', 'CompletedDate', 'DeniedDate', 'InactiveDate', 'PrivacyOptOut',
                   'RentalResolution', 'LMPackageStatus', 'MMSubjectPropertyPresent', 'MMLienInfoPresent'
                                                                                      'LevelOneDate', 'LevelTwoDate',
                   'SeekingShelterResolution', 'YearsAtCurrentAddress'
                                               'SeniorAsHoH', 'HomeOwnerResolutions', 'HomePurchaseResolution']

# Append our proper header to the new worksheet
new_sheet.append(template_header)
new_sheet.title = 'Case'

# Fill our clients hash table with every client
for row in sheet.iter_rows(min_row=4, values_only=True, min_col=2):
    # TODO: IF THE HEADER COLUMNS ARE IN DIFFERENT LOCATIONS, UPDATE THE VALUES FOR THE CLIENT!!!!!
    case = {
        'case_type': row[0],
        'client_id': row[1],
        'assigned_counselor': row[2],
        'assigned_coach': row[3],
        'assigned_loan_officer': row[4],
        'home_purchase_client_type': row[5],
        'home_purchase_client_facilitation': row[6],
        'client_case_status': row[7],
        'client_disclosure_form_present': None,
        'client_first_name': row[9],
        'client_middle_name': row[10],
        'client_last_name': row[11],
        'date_of_birth': correct_date(row[12]),
        'credit_score_before': row[13],
        'credit_score_after': row[14],
        'intake_date': correct_date(row[15]),
        'subsidized_housing_assistance': row[16],
        'primary_employer': row[17],
        'employer_address': row[18],
        'secondary_employer': row[19],
        'secondary_employer_address': row[20],
        'home_owner_last_three_years': row[21],
        'real_estate_agent': row[22],
        'last_contact_date': correct_date(row[23]),
        'long_term_client_date': row[24],
        'short_term_client_date': row[25],
        'near_mortgage_ready_date': row[26],
        'mortgage_ready_date': row[27],
        'in_financing_date': row[28],
        'active_report_date_hud': row[29],
        'completed_date': row[30],
        'denied_date': row[31],
        'inactive_date': row[32],
        'privacy_opt_out': row[33],
        'rental_resolution': row[34],
        'lm_package_status': row[35],
        'mm_subject_property_present': row[36],
        'mm_lien_info_present': row[37],
        'level_one_date': row[38],
        'level_two_date': row[39],
        'seeking_shelter_resolution': row[40],
        'years_at_current_address': row[41],
        'senior_as_hoh': row[42],
        'home_owner_resolutions': row[43],
        'home_purchase_resolution': row[44]
    }
    # Save the client by the ID for easy Access
    cases[case_id] = case
    case_id += 1

# Add each client to the new spreadsheet
for case in cases:
    case_list = [v for k, v in cases[case].items()]

    # # TODO: Fix client data to match our requirements
    # corrected_case = fix_case_data(case_list)

    # Add corrected client to the worksheet
    new_sheet.append(case_list)

# Save the worksheet when all is complete
new_wb.save(filename='gastonia_modified_cases.xlsx')
