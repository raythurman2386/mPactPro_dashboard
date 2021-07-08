from openpyxl import load_workbook
from utils import correct_date, create_workbook
from correct_client import fix_client_data

# Load all files here
client_data = load_workbook(
    filename="MAHA/Contact.xlsx", data_only=True)
client_remaining_data = load_workbook(
    filename="maha_9902.xlsx", data_only=True)
# session_data = load_workbook()
# note_data = load_workbook()

# Create Hash tables for each excel tab
clients = {}
cases = {}
sessions = {}
notes = {}

# *****************************************************************************************
workbook = create_workbook()
template_client_sheet = workbook['Clients']
template_case_sheet = workbook['Case']
template_session_sheet = workbook['Session']
template_note_sheet = workbook['Note']
# *****************************************************************************************

# Start creating initial hash table of clients
for row in client_data.active.iter_rows(min_row=2, values_only=True, min_col=1):
    client_id = row[60]
    client = {
        'clientId': client_id,
        'ClientCaseStatus': None,
        'ClientProgramEnrollment': None,
        'ActiveStaff': None,
        'ClientFirstName': row[5],
        'ClientMiddleName': None,
        'ClientLastName': row[6],
        'DateOfBirth': correct_date(row[36]),
        'Gender': row[84],
        'Race': row[108],
        'Ethnicity': row[88],
        'VeteranStatus': None,
        'ActiveMilitary': None,
        'FirstTimeHomebuyer': None,
        'HouseholdSize': row[86],
        'CountyAmiIncomeLimit': None,
        'HouseholdIncome': row[116],
        'HouseholdIncomeBand': None,
        'IntakeDate': correct_date(row[42]),
        'StreetNumber': None,
        'StreetName': row[16],
        'ApartmentNumber': None,
        'ClientCity': row[17],
        'ClientCounty': None,
        'ClientState': row[18],
        'ClientZip': row[19],
        'PrivacyOptOut': None,
        'RuralAreaStatus': row[109],
        'EnglishProficiencyLevel': row[75],
        'BillToHud': None,
        '8a': None,
        '8b': None,
        '8c': None,
        '8d': None,
        '8e': None,
        '8f': None,
        '8g': None,
        '8h': None,
        '8i': None,
        '9a': None,
        '9b': None,
        '9c': None,
        '9d': None,
        '9e': None,
        '9f': None,
        '10a': None,
        '10b': None,
        '10c': None,
        '10d': None,
        '10e': None,
        '10f': None,
        '10g': None,
        '10h': None,
        '10i': None,
        '10j': None,
        '10k': None,
        '10l': None,
        '10m': None,
        'PhoneNumberMobile': row[26],
        'PhoneNumberWork': None,
        'PhoneNumberHome': row[27],
        'ImmigrationStatus': None,
        'EmailHome': row[31],
        'EmailWork': row[31],
        'MaritalStatus': row[92],
        'Disability': row[71],
        'HouseholdType': row[85],
        'Education': row[74],
        'ReferralSource': row[61],
        'LastContact': None,
        'ActiveReportDateHUD': None,
        'CompletedDate': None,
        'InactiveDate': None
    }

    # Save the client by the ID for easy Access
    clients[client_id] = client

# THIS IS DONE LAST FOR THE CLIENT SHEET
# GRAB ALL DATA BEFORE THIS UNLESS TESTING
# Add each client to the new spreadsheet
for client in clients:
    # Need to pull over the address correction from the main file
    client_list = [v for k, v in clients[client].items()]

    # If the address needs split, pass True along with the client list
    # If correcting address use this variable instead
    corrected_client = fix_client_data(client_list, True)
    # If the address does not need split this is the one to use
    # corrected_client = fix_client_data(client_list)

    template_client_sheet.append(client_list)

# *****************************************************************************************
# Start creating initial hash table of cases
# for row in _.active.iter_rows(min_row=2, values_only=True, min_col=1):
#     pass

# *****************************************************************************************
# Start creating initial hash table of sessions
# for row in _.active.iter_rows(min_row=2, values_only=True, min_col=1):
#     pass

# *****************************************************************************************
# Start creating initial hash table of notes
# for row in client_notes.active.iter_rows(min_row=2, values_only=True, min_col=1):
#     # TODO: IF THE HEADER COLUMNS ARE IN DIFFERENT LOCATIONS, UPDATE THE VALUES FOR THE CLIENT!!!!!
#     note_id = row[0]
#     client_id = row[1]
#     note = {
#         'client_id': client_id,
#         'first_name': clients[client_id]["ClientFirstName"],
#         'last_name': clients[client_id]["ClientLastName"],
#         'duration': None,
#         'date_start': row[2],
#         'counselor_name': None,
#         'notes': row[3],
#     }
#     # Save the client by the ID for easy Access
#     notes[note_id] = note


# # Add each client to the new spreadsheet
# for note in notes:
#     note_list = [v for k, v in notes[note].items()]
#     # Add corrected client to the worksheet
#     template_note_sheet.append(note_list)
# *****************************************************************************************


# Save the worksheet when all is complete
outputFileName = "test_maha.xlsx"
workbook.save(filename=outputFileName)
