import datetime
from openpyxl import load_workbook, Workbook


def correct_date(date):
    try:
        if date is not None and date is not str:
            return date.strftime('%m/%d/%Y')
    except AttributeError:
        return date


def clean_phone(phone):
    phone_str = str(phone)
    try:
        if phone is not None:
            if '-' in phone_str and '(' not in phone_str and ')' not in phone_str:
                new_num = phone_str.replace('-', '')
                phone = int(new_num)
            elif '(' in phone_str and ')' in phone_str and '-' in phone_str:
                phone_num = phone_str.replace('(', '')
                phone_num2 = phone_num.replace(')', '')
                new_num = phone_num2.replace('-', '')
                phone = int(new_num)
            elif '(' in phone_str and ')' in phone_str:
                phone_num = phone_str.replace('(', '')
                phone_num2 = phone_num.replace(')', '')
                phone = int(phone_num2)
    except ValueError:
        pass

    return phone


def get_phone(client, a, b):
    if client[a] is None:
        return clean_phone(client[b])

    return clean_phone(client[a])


def check_grant(row, date):
    if row is not None:
        return date

    return


def update_client_fields(client, nine, date):
    new_fields = {
        'CaseID': client[4],
        'Date': date,
        'NOFAGrant': client[38],
        'Zip': client[9],
        nine: date,
        '10a': check_grant(client[40], date),
        '10b': check_grant(client[41], date),
        '10c': check_grant(client[42], date),
        '10d': check_grant(client[43], date),
        '10e': check_grant(client[44], date),
        '10f': check_grant(client[45], date),
        '10i': check_grant(client[46], date),
        '10j': check_grant(client[47], date),
        '10l': check_grant(client[48], date),
        '10m': check_grant(client[49], date),
    }

    return new_fields


def create_workbook():
    # *****************************************************************************************
    # Open new workbook, create tabs, and write headers
    # Full template workbook setup and initialization
    workbook = Workbook()
    template_client_sheet = workbook.create_sheet('Client Case Session Upload')
    template_case_sheet = workbook.create_sheet('Client Class Upload')

    # Headers for all tabs of the workbook
    client_case_session_upload_header = ['ClientID', 'FirstName', 'LastName', 'Phone', 'AddressNumber', 'StreetName', 'City', 'State', 'County',
                                         'Zip', 'Race', 'Ethnicity', 'EnglishProficiencyLevel', 'HouseholdIncome',
                                         'HouseholdSize', 'ApartmentNumber', 'Email', 'InitialCaseType', 'DateofBirth', 'CounselorName', 'IntakeDate', 'CaseType', 'CaseStartDate', 'CaseID',
                                         'SessionID', 'Date', 'NOFAGrant', '9a', '9b', '9c', '9d', '9e', '9f', '10a', '10b', '10c', '10d',
                                         '10e', '10f', '10g', '10h', '10i', '10j', '10k', '10l', '10m', '10n', 'SessionNotes']
    client_class_upload_header = ['ClientID', 'FirstName', 'LastName', 'Phone', 'AddressNumber', 'StreetName', 'City', 'State', 'County',
                                  'Zip', 'Race', 'Ethnicity', 'EnglishProficiencyLevel', 'HouseholdIncome',
                                  'HouseholdSize', 'ApartmentNumber', 'Email', 'InitialCaseType', 'DateofBirth', 'CounselorName', 'IntakeDate',
                                  'CourseID', 'ClassDate', 'AttendanceStatus']

    # Append our proper header to the new worksheet
    template_client_sheet.append(client_case_session_upload_header)
    template_case_sheet.append(client_class_upload_header)
    # *****************************************************************************************
    workbook.remove(workbook.active)
    return workbook


options = {
    'race': {
        'a': 'a. American Indian/Alaskan Native',
        'b': 'b. Asian',
        'c': 'c. Black or African American',
        'd': 'd. Native Hawaiian or Other Pacific Islander',
        'e': 'e. White',
        'f': 'f. American Indian or Alaska Native and White',
        'g': 'g. Asian and White',
        'h': 'h. Black or African American and White',
        'i': 'i. American Indian or Alaska Native and Black or African American',
        'j': 'j. Other multiple race',
        'k': 'k. Chose not to respond'},
    'ethnicity': {
        'a': 'a. Hispanic',
        'b': 'b. Not Hispanic',
        'c': 'c. Chose not to respond'},
    'englishProficiency': {
        'a': 'a. Household is Limited English Proficient',
        'b': 'b . Household is not Limited English Proficient',
        'c': 'c. Chose not to respond'
    },
    'caseType': {
        'a': 'Home Purchase',
        'b': 'Seeking Shelter',
        'c': 'Home Owner Services',
        'd': 'Mortgage Modification',
        'e': 'Rental topics'
    },
    'NOFA': {
        'a': 'NOFA 2017-1 COMP',
        'b': 'NOFA 2019-1 COMP',
        'c': 'NOFA 2020-1 COMP',
        'd': 'NOFA 2021-1 COMP'
    },
    'Attended': {
        'a': 'Attended',
        'b': 'No Show',
        'c': 'Enrolled'
    }
}
