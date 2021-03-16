from openpyxl import load_workbook
from utils import correct_date

# This is the file that will be corrected
# masterFile = sys.argv[1]

# This is the file that will be getting corrected
# file = load_workbook(filename=masterFile, data_only=True)

# case_sheet = file.active

# This will be a new workbook that we save our corrected values to
workbook = load_workbook(filename='AccessLiving.xlsx', data_only=True)
new_wb = load_workbook(filename='AccessLiving_modified.xlsx')
sheet = workbook['Notes']
new_sheet = new_wb.create_sheet('Note')

# Create a hashtable to keep track of cases
notes = {}
note_id = 1

template_header = ['Client ID', 'First Name', 'Last Name', 'Duration', 'Date - start', 'Counselor - name', 'Notes']

# Fill our clients hash table with every client
for row in sheet.iter_rows(min_row=2, values_only=True, min_col=1):
    # TODO: IF THE HEADER COLUMNS ARE IN DIFFERENT LOCATIONS, UPDATE THE VALUES FOR THE CLIENT!!!!!
    note = {
        'client_id': row[1],
        'first_name': None,
        'last_name': None,
        'duration': None,
        'date_start': row[2],
        'counselor_name': None,
        'notes': row[3],
    }
    # Save the client by the ID for easy Access
    notes[note_id] = note
    note_id += 1

# Append our proper header to the new worksheet
new_sheet.append(template_header)
new_sheet.title = 'Note'

# Add each client to the new spreadsheet
for note in notes:
    note_list = [v for k, v in notes[note].items()]

    # Add corrected client to the worksheet
    new_sheet.append(note_list)

# Save the worksheet when all is complete
new_wb.save(filename='gastonia_modified_notes.xlsx')
