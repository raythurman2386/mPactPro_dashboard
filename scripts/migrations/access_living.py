# Specific file where all of Access Livings data will be manipulated.
# Open all needed files
# Grab correct columns from every file to fill our template
# Create a hash table of clients using ClientID as the Hash table Key
# (Columns with options will be corrected after populating our template)
# Create a hash table off of the dropdown options file using the option for the Hash table Key
# Correct columns that are using options by utilizing the created hash table
# Once all columns are populated then the regular process of updating the data to match our points starts.

from openpyxl import load_workbook, Workbook
import sys
import os

# Load all files here
client_data = load_workbook(filename="access living/client.xlsx", data_only=True)
client_address = load_workbook(filename="access living/client address.xlsx", data_only=True)
client_notes = load_workbook(filename="access living/client notes.xlsx", data_only=True)
client_phone = load_workbook(filename="access living/client phone.xlsx", data_only=True)
person = load_workbook(filename="access living/person.xlsx", data_only=True)
housing_info = load_workbook(filename="access living/housing information.xlsx", data_only=True)
demographics = load_workbook(filename="access living/demographic.xlsx", data_only=True)

# Dropdown code files
dropdown_codes = load_workbook(filename="access living/dropdown codes.xlsx", data_only=True)
housing_codes = load_workbook(filename="access living/housing dropdown codes.xlsx", data_only=True)

# Create Hash tables for each excel tab
clients = {}
cases = {}
sessions = {}
notes = {}

# Open new workbook and assign vars for sheet tabs
# This will be a new workbook that we save our corrected values to
workbook = Workbook()
template_client_sheet = workbook.active
# template_case_sheet = workbook['Case']
# template_session_sheet = workbook['Session']
# template_note_sheet = workbook['Note']

# Headers for all tabs of the workbook
client_header = ['ClientID', 'ClientCaseStatus', 'ClientProgramEnrollment', 'ActiveStaff', 'ClientFirstName',
                 'ClientMiddleName',
                 'ClientLastName', 'DateOfBirth', 'Gender', 'Race', 'Ethnicity', 'VeteranStatus', 'ActiveMilitary',
                 'FirstTimeHomebuyer', 'HouseholdSize', 'CountyAmiIncomeLimit', 'HouseholdIncome',
                 'HouseholdIncomeBand',
                 'IntakeDate', 'StreetNumber', 'StreetName', 'ApartmentNumber', 'ClientCity', 'ClientCounty',
                 'ClientState',
                 'ClientZip', 'PrivacyOptOut', 'RuralAreaStatus', 'EnglishProficiencyLevel', 'BillToHud', '8a', '8b',
                 '8c', '8d',
                 '8e', '8f', '8g', '8h', '8i', '9a', '9b', '9c', '9d', '9e', '9f', '10a', '10b', '10c', '10d', '10e',
                 '10f',
                 '10g', '10h', '10i', '10j', '10k', '10l', '10m', 'PhoneNumberMobile', 'PhoneNumberWork',
                 'PhoneNumberHome',
                 'ImmigrationStatus', 'EmailHome', 'EmailWork', 'MaritalStatus', 'Disability', 'HouseholdType',
                 'Education',
                 'ReferralSource', 'LastContact', 'ActiveReportDateHUD', 'CompletedDate', 'InactiveDate']
case_header = ['CaseType', 'ClientID', 'AssignedCounselor', 'AssignedCoach', 'AssignedLoanOfficer',
               'HomePurchaseClientType', 'HomePurchaseClientFacilitation', 'ClientCaseStatus',
               'ClientDisclosureFormPresent', 'ClientFirstName', 'ClientMiddleName', 'ClientLastName',
               'DateofBirth', 'CreditScoreBefore', 'CreditScoreAfter', 'IntakeDate', 'SubsidizedHousingAssistance',
               'PrimaryEmployer', 'EmployerAddress', 'SecondaryEmployer', 'SecondaryEmployerAddress',
               'HomeOwnerLastThreeYears', 'RealEstateAgent', 'LastContactDate', 'LongTermClientDate',
               'ShortTermClientDate', 'NearMortgageReadyDate', 'MortgageReadyDate', 'InFinancingDate',
               'ActiveReportDateHUD', 'CompletedDate', 'DeniedDate', 'InactiveDate', 'PrivacyOptOut',
               'RentalResolution', 'LMPackageStatus', 'MMSubjectPropertyPresent', 'MMLienInfoPresent'
                                                                                  'LevelOneDate', 'LevelTwoDate',
               'SeekingShelterResolution', 'YearsAtCurrentAddress'
                                           'SeniorAsHoH', 'HomeOwnerResolutions', 'HomePurchaseResolution']
session_header = ['SessionDuration', 'CounselorName', 'SessionType', 'ClientNotes', '9series', '10a', '10b', '10c',
                  '10d', '10e', '10f', '10g', '10h', '10i', '10j', '10k', '10l', '10m', 'SessionFee', 'BillableTo',
                  'IfOtherWho']
note_header = ['ClientID', 'ClientFirstName', 'ClientLastName', 'Duration', 'NoteDate', 'Counselor', 'NoteText']

# Start creating initial hash table of clients from the PERSON workbook

first = person.active
for row in first.iter_rows(min_row=2, values_only=True, min_col=1):
    client_id = row[0]
    client = {
        'clientId': client_id,
        'ClientCaseStatus': None,
        'ClientProgramEnrollment': None,
        'ActiveStaff': None,
        'ClientFirstName': row[2],
        'ClientMiddleName': row[3],
        'ClientLastName': row[4],
        'DateOfBirth': row[7],
        'Gender': row[6],
        'Race': None,
        'Ethnicity': None,
        'VeteranStatus': None,
        'ActiveMilitary': None,
        'FirstTimeHomebuyer': None,
        'HouseholdSize': None,
        'CountyAmiIncomeLimit': None,
        'HouseholdIncome': None,
        'HouseholdIncomeBand': None,
        'IntakeDate': row[12],
        'StreetNumber': None,
        'StreetName': None,
        'ApartmentNumber': None,
        'ClientCity': None,
        'ClientCounty': None,
        'ClientState': None,
        'ClientZip': None,
        'PrivacyOptOut': None,
        'RuralAreaStatus': None,
        'EnglishProficiencyLevel': None,
        'BillToHud': None,
        '8a': None,
        '8b': None,
        '8c': None,
        '8d': None,
        '8e': None,
        '8f': None,
        '8g': None,
        '8h': None,
        '8i': None,
        '9a': None,
        '9b': None,
        '9c': None,
        '9d': None,
        '9e': None,
        '9f': None,
        '10a': None,
        '10b': None,
        '10c': None,
        '10d': None,
        '10e': None,
        '10f': None,
        '10g': None,
        '10h': None,
        '10i': None,
        '10j': None,
        '10k': None,
        '10l': None,
        '10m': None,
        'PhoneNumberMobile': None,
        'PhoneNumberWork': None,
        'PhoneNumberHome': None,
        'ImmigrationStatus': None,
        'EmailHome': row[9],
        'EmailWork': row[10],
        'MaritalStatus': None,
        'Disability': None,
        'HouseholdType': None,
        'Education': None,
        'ReferralSource': None,
        'LastContact': None,
        'ActiveReportDateHUD': None,
        'CompletedDate': None,
        'InactiveDate': None
    }

    # Save the client by the ID for easy Access
    clients[client_id] = client

# Now that initial clients hashtable is created, go through each open file and fill in client data
second = client_data.active
for row in second.iter_rows(min_row=2, min_col=1, values_only=True):
    client_id = row[0]
    try:
        if clients[client_id]:
            clients[client_id]["MaritalStatus"] = row[2]
            clients[client_id]["ReferralSource"] = row[1]
            clients[client_id]["HouseholdSize"] = row[5]
            clients[client_id]["Ethnicity"] = row[6]
    except KeyError:
        pass

# Append our proper header to the new worksheet
template_client_sheet.append(client_header)
template_client_sheet.title = 'Clients'

# Add each client to the new spreadsheet
for client in clients:
    client_list = [v for k, v in clients[client].items()]
    template_client_sheet.append(client_list)

# Save the worksheet when all is complete
outputFileName = "test_modified.xlsx"
workbook.save(filename=outputFileName)