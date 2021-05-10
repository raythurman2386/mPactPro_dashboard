import datetime
from openpyxl import load_workbook, Workbook


def correct_date(date):
    try:
        if date is not None and date is not str:
            return date.strftime('%m/%d/%Y')
    except AttributeError:
        return date


def create_workbook():
    # *****************************************************************************************
    # Open new workbook, create tabs, and write headers
    # Full template workbook setup and initialization
    workbook = Workbook()
    template_client_sheet = workbook.create_sheet('Clients')
    template_case_sheet = workbook.create_sheet('Case')
    template_session_sheet = workbook.create_sheet('Session')
    template_note_sheet = workbook.create_sheet('Note')

    # Headers for all tabs of the workbook
    client_header = ['ClientID', 'ClientCaseStatus', 'ClientProgramEnrollment', 'ActiveStaff', 'ClientFirstName',
                     'ClientMiddleName',
                     'ClientLastName', 'DateOfBirth', 'Gender', 'Race', 'Ethnicity', 'VeteranStatus', 'ActiveMilitary',
                     'FirstTimeHomebuyer', 'HouseholdSize', 'CountyAmiIncomeLimit', 'HouseholdIncome',
                     'HouseholdIncomeBand',
                     'IntakeDate', 'StreetNumber', 'StreetName', 'ApartmentNumber', 'ClientCity', 'ClientCounty',
                     'ClientState',
                     'ClientZip', 'PrivacyOptOut', 'RuralAreaStatus', 'EnglishProficiencyLevel', 'BillToHud', '8a',
                     '8b',
                     '8c', '8d',
                     '8e', '8f', '8g', '8h', '8i', '9a', '9b', '9c', '9d', '9e', '9f', '10a', '10b', '10c', '10d',
                     '10e',
                     '10f',
                     '10g', '10h', '10i', '10j', '10k', '10l', '10m', 'PhoneNumberMobile', 'PhoneNumberWork',
                     'PhoneNumberHome',
                     'ImmigrationStatus', 'EmailHome', 'EmailWork', 'MaritalStatus', 'Disability', 'HouseholdType',
                     'Education',
                     'ReferralSource', 'LastContact', 'ActiveReportDateHUD', 'CompletedDate', 'InactiveDate']
    case_header = ['CaseType', 'ClientID', 'AssignedCounselor', 'AssignedCoach', 'AssignedLoanOfficer',
                   'HomePurchaseClientType', 'HomePurchaseClientFacilitation', 'ClientCaseStatus',
                   'ClientDisclosureFormPresent', 'ClientFirstName', 'ClientMiddleName', 'ClientLastName',
                   'DateofBirth', 'CreditScoreBefore', 'CreditScoreAfter', 'IntakeDate', 'SubsidizedHousingAssistance',
                   'PrimaryEmployer', 'EmployerAddress', 'SecondaryEmployer', 'SecondaryEmployerAddress',
                   'HomeOwnerLastThreeYears', 'RealEstateAgent', 'LastContactDate', 'LongTermClientDate',
                   'ShortTermClientDate', 'NearMortgageReadyDate', 'MortgageReadyDate', 'InFinancingDate',
                   'ActiveReportDateHUD', 'CompletedDate', 'DeniedDate', 'InactiveDate', 'PrivacyOptOut',
                   'RentalResolution', 'LMPackageStatus', 'MMSubjectPropertyPresent', 'MMLienInfoPresent'
                                                                                      'LevelOneDate', 'LevelTwoDate',
                   'SeekingShelterResolution', 'YearsAtCurrentAddress'
                                               'SeniorAsHoH', 'HomeOwnerResolutions', 'HomePurchaseResolution']
    session_header = ['SessionID', 'ClientID', 'TimeDateSession', 'SessionDuration', 'CounselorName', 'SessionType', 'ClientNotes', '9series', '10a', '10b', '10c',
                      '10d', '10e', '10f', '10g', '10h', '10i', '10j', '10k', '10l', '10m', 'SessionFee', 'BillableTo',
                      'IfOtherWho']
    note_header = ['ClientID', 'ClientFirstName', 'ClientLastName',
                   'Duration', 'NoteDate', 'Counselor', 'NoteText']

    # Append our proper header to the new worksheet
    template_client_sheet.append(client_header)
    template_case_sheet.append(case_header)
    template_session_sheet.append(session_header)
    template_note_sheet.append(note_header)
    # *****************************************************************************************
    workbook.remove(workbook.active)
    return workbook


options = {
    'yes_no': {
        'a': 'Yes',
        'b': 'No'
    },
    'ClientCaseStatus': {
        'a': 'New Intake',
        'b': 'Active',
        'c': 'Completed',
        'd': 'Withdrew',
        'e': 'Inactive',
        'f': 'Completed - Education Only'
    },
    'ClientProgramEnrollment': {
        'a': 'Home Purchase',
        'b': 'Mortgage Modification',
        'c': 'Homelessness Assistance',
        'd': 'Rental Topics',
        'e': 'Homeowner Services',
        'f': 'Education Services'
    },
    'gender': ['Male', 'Female'],
    'race': {
        'a': 'a. American Indian/Alaskan Native',
        'b': 'b. Asian',
        'c': 'c. Black or African American',
        'd': 'd. Native Hawaiian or Other Pacific Islander',
        'e': 'e. White',
        'f': 'f. American Indian or Alaska Native and White',
        'g': 'g. Asian and White',
        'h': 'h. Black or African American and White',
        'i': 'i. American Indian or Alaska Native and Black or African American',
        'j': 'j. Other multiple race',
        'k': 'k. Chose not to respond'},
    'ethnicity': {
        'a': 'a. Hispanic',
        'b': 'b. Not Hispanic',
        'c': 'c. Chose not to respond'},
    'incomeBand': {
        'a': '1. Below 30% of AMI',
        'b': '2. 30% - 49% of AMI',
        'c': '3. 50% - 79% of AMI',
        'd': '4. 80% - 100% of AMI',
        'e': '5. 101% - 120% of AMI',
        'f': '6. Chose not to respond'
    },
    'ruralStatus': {
        'a': 'a. Household lives in a rural area',
        'b': 'b. Household does not live in a rural area',
        'c': 'c. Chose not to respond'
    },
    'englishProficiency': {
        'a': 'a. Household is Limited English Proficient',
        'b': 'b . Household is not Limited English Proficient',
        'c': 'c. Chose not to respond'
    },
    'maritalStatus': {
        'a': 'Single',
        'b': 'Married',
        'c': 'Divorced',
        'd': 'Seperated',
        'e': 'Widowed',
    },
    'householdType': {
        'a': '1. Female headed single parent household',
        'b': '2. Male headed single parent household',
        'c': '3. Single adult',
        'd': '4. Two or more unrelated adults',
        'e': '5. Married with children',
        'f': '6. Married without children',
        'g': '7. Other',
    },
    'education': {
        'a': '1. Below High School Diploma',
        'b': '2. High School Diploma or Equivalent',
        'c': '3. Two-Year College',
        'd': '4. Bachelors Degree',
        'e': '5. Masters Degree',
        'f': '6. Above Masters Degree',
        'g': 'Unknown',
        'h': 'Other'
    },
    'referralSource': {
        'a': '1. Print Advertisement',
        'b': '2. Bank',
        'c': '3. Government',
        'd': '4. TV',
        'e': '5. Realtor',
        'f': '6. Staff/Board Member',
        'g': '7. Walk-in',
        'h': '8. Friend',
        'i': '9. Radio',
        'j': '10. Newspaper Article',
        'k': '11. Other',
    }
}


case_options = {
    'caseType': {
        'a': 'Home Purchase',
        'b': 'Seeking Shelter',
        'c': 'Home Owner Services',
        'd': 'Mortgage Modification',
        'e': 'Rental topics'
    },
    'homePurchaseClientType': {
        'a': 'Pre Purchase Counseling Client',
        'b': 'Pre Purchase Education Client',
        'c': 'eHome America Education Client',
        'd': 'Financial Literacy Client'
    },
    'homePurchaseFacilitation': {
        'a': 'Financial Literacy',
        'b': 'Home Purchase'
    },
    'clientCaseStatus': {
        'a': 'Home Purchase Case',
        'b': 'Long Term Client',
        'c': 'Short Term Client',
        'd': 'Near Mortgage Ready',
        'e': 'Mortgage Ready',
        'f': 'Internal Financing',
        'g': 'External Financing',
        'h': 'Denied',
        'i': 'Withdrew',
        'j': 'Inactive',
        'k': 'Completed',
        'l': 'New Mortgage Modification Case',
        'm': 'Level 1 Client',
        'n': 'Level 2 Client',
        'o': 'Homelessness Assistance Case',
        'p': 'Rental Topics Case'
    },
    'yes_no': {
        'a': 'Yes',
        'b': 'No'
    },
    'rentalResolution': {
        'a': 'Found Alternative',
        'b': 'Suspended',
        'c': 'Decided to Remain in Current Housing',
        'd': 'Received Housing Search Assistance',
        'e': 'Referred to Other Agency w/ Rental Assistance',
        'f': 'Referred to Other Agency w/ Social Assistance',
        'g': 'Advised on Re-Cert for HUD',
        'h': 'Counseled/Referred to Legal Aid',
        'i': 'Entered Debt Management/Repayment Plan',
        'j': 'Temp Rental Relief',
        'k': 'Counseled and Utilities Brought Current',
        'l': 'Resolved Security Deposit Dispute',
        'm': 'Resolved issue in Current Tenacy'
    },
    'lmPackageStatus': {
        'a': 'New',
        'b': 'In Progress',
        'c': 'Complete',
        'd': 'Incomplete'
    },
    'seekingShelterResolution': {
        'a': 'Suspended',
        'b': 'Emergency Shelter',
        'c': 'Transitional Housing',
        'd': 'Remained in Housing',
        'e': 'Remained Homeless',
        'f': 'Permanent Housing w/Rental Assistance',
        'g': 'Permanent Housing w/out Rental Assistance',
        'h': 'Referred to Other Social Service Agency'
    },
    'homeOwnerResolutions': {
        'a': 'Counseled On HECM; Declined',
        'b': 'Referred To other Agency',
        'c': 'Obtained HECM',
        'd': 'Received Home Equity/Home Improvement',
        'e': 'Mortgage Refinanced',
        'f': 'Sold House, Chose Alt. Housing Solution',
        'g': 'Suspended',
        'h': 'Received Consumer Loan (Unsecured)',
        'i': 'Obtained a Non-FHA Reverse Mortgage Loan',
        'j': 'Completed Financial Management/Budget Counseling',
        'k': 'Complete Home Maintenance Counseling',
        'l': 'Counseled and Utilities Brought Current',
        'm': 'Counseled/Referred for Legal Assistance'
    },
    'homePurchaseResolution': {
        'a': 'New File',
        'b': 'Closed-Financial Literacy',
        'c': 'Open Case',
        'd': 'Closed-Home Purchased'
    }
}


session_options = {
    'sessionType': {
        'a': 'General Session',
        'b': 'General Session',
        'c': 'Client Did Not Show',
        'd': 'Level 1 Session',
        'e': 'Level 2 Session',
        'f': 'Level 1 & 2 Session',
        'g': 'Long Term - Pre-Purchase Home Buyer Counseling',
        'h': 'Short Term - Pre-Purchase Home Buyer Counseling',
        'i': 'Near Mortgage Ready - Pre-Purchase Home Buyer Counseling',
        'j': 'Mortgage Ready - Pre-Purchase Home Buyer Counseling',
        'k': 'Internal Closing Counseling Session',
        'l': 'General Coaching',
        'm': 'Education Follow-Up',
        'n': 'Financial management/ budget counseling',
        'o': 'Homelessness Assistance Counseling',
        'p': 'Rental Topics Work Session',
        'q': 'Homeowner Services Work Session',
        'r': 'Reverse Mortgage Counseling'
    },
    'billableTo': {
        'a': 'Non-Billable',
        'b': 'HUD',
        'c': 'Neighborworks',
        'd': 'Other'
    }
}
