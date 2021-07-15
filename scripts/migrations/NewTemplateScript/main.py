from openpyxl import load_workbook
from utils import correct_date, create_workbook
from correct_data import correct_data

# Load all files here
client_data = load_workbook(
    filename="Contact.xlsx", data_only=True)
client_9902_data = load_workbook(
    filename="maha_9902.xlsx", data_only=True)
# session_data = load_workbook()
# note_data = load_workbook()

# Create Hash tables for each excel tab
clients = {}
classes = {}

# *****************************************************************************************
workbook = create_workbook()
template_client_sheet = workbook['Client Case Session Upload']
template_class_sheet = workbook['Client Class Upload']
# *****************************************************************************************

# Start creating initial hash table of clients
for row in client_data.active.iter_rows(min_row=2, values_only=True, min_col=1):
    client_id = row[60]
    client = {
        'clientId': client_id,
        'FirstName': row[5],
        'LastName': row[6],
        'Phone': None,
        'AddressNumber': None,
        'StreetName': row[16],
        'ClientCity': row[17],
        'ClientState': row[18],
        'ClientCounty': None,
        'Zip': row[19],
        'Race': row[108],
        'Ethnicity': row[88],
        'EnglishProficiencyLevel': row[75],
        'HouseholdIncome': row[116],
        'HouseholdSize': row[86],
        'ApartmentNumber': None,
        'Email': row[31],
        'InitialCaseType': None,
        'DateOfBirth': correct_date(row[36]),
        'CounselorName': None,
        'IntakeDate': correct_date(row[42]),
        'CaseType': None,
        'CaseStartDate': correct_date(row[42]),
        'CaseID': None,
        'SessionID': None,
        'Date': None,
        'NOFAGrant': None,
        '9a': None,
        '9b': None,
        '9c': None,
        '9d': None,
        '9e': None,
        '9f': None,
        '10a': None,
        '10b': None,
        '10c': None,
        '10d': None,
        '10e': None,
        '10f': None,
        '10g': None,
        '10h': None,
        '10i': None,
        '10j': None,
        '10k': None,
        '10l': None,
        '10m': None,
        '10n': None,
        'SessionNotes': None
    }

    # Save the client by the ID for easy Access
    clients[client_id] = client

# THIS IS DONE LAST FOR THE CLIENT SHEET
# GRAB ALL DATA BEFORE THIS UNLESS TESTING
# Add each client to the new spreadsheet
for client in clients:
    # Need to pull over the address correction from the main file
    client_list = [v for k, v in clients[client].items()]

    # If the address needs split, pass True along with the client list
    # If correcting address use this variable instead
    corrected_client = correct_data(client_list, True, 'Client_Case')
    # If the address does not need split this is the one to use
    # corrected_client = fix_client_data(client_list, 'Client_Case')

    template_client_sheet.append(client_list)

# *****************************************************************************************
# Start creating initial hash table of classes
for row in client_data.active.iter_rows(min_row=2, values_only=True, min_col=1):
    client_id = row[60]
    client = {
        'clientId': client_id,
        'FirstName': row[5],
        'LastName': row[6],
        'Phone': None,
        'AddressNumber': None,
        'StreetName': row[16],
        'ClientCity': row[17],
        'ClientState': row[18],
        'ClientCounty': None,
        'Zip': row[19],
        'Race': row[108],
        'Ethnicity': row[88],
        'EnglishProficiencyLevel': row[75],
        'HouseholdIncome': row[116],
        'HouseholdSize': row[86],
        'ApartmentNumber': None,
        'Email': row[31],
        'InitialCaseType': None,
        'DateOfBirth': correct_date(row[36]),
        'CounselorName': None,
        'IntakeDate': correct_date(row[42]),
        'CourseID': None,
        'ClassDate': None,
        'AttendanceStatus': None
    }

    # Save the client by the ID for easy Access
    classes[client_id] = client

for client in classes:
    # Need to pull over the address correction from the main file
    client_list = [v for k, v in classes[client].items()]

    # If the address needs split, pass True along with the client list
    # If correcting address use this variable instead
    corrected_client = correct_data(client_list, True, 'Client_Class')
    # If the address does not need split this is the one to use
    # corrected_client = fix_client_data(client_list, 'Client_Class')

    template_class_sheet.append(client_list)


# Fill in remaining Case Session or Class Data from 9902 File
for row in client_9902_data.active.iter_rows(min_row=11, min_col=4, values_only=True):
    client_id = row[1]
    try:
        clients[client_id]['CaseID'] = row[2]
        clients[client_id]['Date'] = correct_date(row[24])
        clients[client_id]['NOFAGrant'] = row[36]
    except KeyError:
        print('Key Error')


# Save the worksheet when all is complete
outputFileName = "new_template_maha.xlsx"
workbook.save(filename=outputFileName)
