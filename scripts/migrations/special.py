from openpyxl import load_workbook, Workbook


# data_only flag will only bring in the data and no excel functions
file = load_workbook(filename='AccessLiving.xlsx', data_only=True)

client_sheet = file['Clients']

# This will be a new workbook that we save our corrected values to
workbook = Workbook()
sheet = workbook.active

# Create a hashtable to keep track of clients
clients = {}

template_header = ['ClientID', 'ClientCaseStatus', 'ClientProgramEnrollment', 'ActiveStaff', 'ClientFirstName',
                   'ClientMiddleName',
                   'ClientLastName', 'DateOfBirth', 'Gender', 'Race', 'Ethnicity', 'VeteranStatus', 'ActiveMilitary',
                   'FirstTimeHomebuyer', 'HouseholdSize', 'CountyAmiIncomeLimit', 'HouseholdIncome',
                   'HouseholdIncomeBand',
                   'IntakeDate', 'StreetNumber', 'StreetName', 'ApartmentNumber', 'ClientCity', 'ClientCounty',
                   'ClientState',
                   'ClientZip', 'PrivacyOptOut', 'RuralAreaStatus', 'EnglishProficiencyLevel', 'BillToHud', '8a', '8b',
                   '8c', '8d',
                   '8e', '8f', '8g', '8h', '8i', '9a', '9b', '9c', '9d', '9e', '9f', '10a', '10b', '10c', '10d', '10e',
                   '10f',
                   '10g', '10h', '10i', '10j', '10k', '10l', '10m', 'PhoneNumberMobile', 'PhoneNumberWork',
                   'PhoneNumberHome',
                   'ImmigrationStatus', 'EmailHome', 'EmailWork', 'MaritalStatus', 'Disability', 'HouseholdType',
                   'Education',
                   'ReferralSource', 'LastContact', 'ActiveReportDateHUD', 'CompletedDate', 'InactiveDate', 'item_ID',
                   'Source']


# Fill our clients hash table with every client
for row in client_sheet.iter_rows(min_row=2, values_only=True, min_col=1):
    client_id = row[0]
    # TODO: IF THE HEADER COLUMNS ARE IN DIFFERENT LOCATIONS, UPDATE THE VALUES FOR THE CLIENT!!!!!
    # This client matches our Template DO NOT CHANGE end up creating multiple
    # clients for each different cms we are importing.
    client = {
      'clientId': client_id,
      'ClientCaseStatus': None,
      'ClientProgramEnrollment': None,
      'ActiveStaff': None,
      'ClientFirstName': row[2],
      'ClientMiddleName': row[3],
      'ClientLastName': row[4],
      'DateOfBirth': row[7],
      'Gender': row[6],
      'Race': None,
      'Ethnicity': None,
      'VeteranStatus': None,
      'ActiveMilitary': None,
      'FirstTimeHomebuyer': None,
      'HouseholdSize': None,
      'CountyAmiIncomeLimit': None,
      'HouseholdIncome': None,
      'HouseholdIncomeBand': None,
      'IntakeDate': row[12],
      'StreetNumber': None,
      'StreetName': None,
      'ApartmentNumber': None,
      'ClientCity': None,
      'ClientCounty': None,
      'ClientState': None,
      'ClientZip': None,
      'PrivacyOptOut': None,
      'RuralAreaStatus': None,
      'EnglishProficiencyLevel': None,
      'BillToHud': None,
      '8a': None,
      '8b': None,
      '8c': None,
      '8d': None,
      '8e': None,
      '8f': None,
      '8g': None,
      '8h': None,
      '8i': None,
      '9a': None,
      '9b': None,
      '9c': None,
      '9d': None,
      '9e': None,
      '9f': None,
      '10a': None,
      '10b': None,
      '10c': None,
      '10d': None,
      '10e': None,
      '10f': None,
      '10g': None,
      '10h': None,
      '10i': None,
      '10j': None,
      '10k': None,
      '10l': None,
      '10m': None,
      'PhoneNumberMobile': None,
      'PhoneNumberWork': None,
      'PhoneNumberHome': None,
      'ImmigrationStatus': None,
      'EmailHome': row[9],
      'EmailWork': row[10],
      'MaritalStatus': None,
      'Disability': None,
      'HouseholdType': None,
      'Education': None,
      'ReferralSource': None,
      'LastContact': None,
      'ActiveReportDateHUD': None,
      'CompletedDate': None,
      'InactiveDate': None,
      'Item_id': None,
      'Source': None
    }
    clients[client_id] = client

# Check other Tabs in original workbook for additional client data
# Match Client ID with Client in the hash table and update missing fields.
client_address = file['Address']
for row in client_address.iter_rows(min_row=2, values_only=True, min_col=1):
    try:
        client_id_2 = row[0]
        if clients[client_id_2]:
            clients[client_id_2]['StreetName'] = row[5]
            clients[client_id_2]['ClientCounty'] = row[6]
            clients[client_id_2]['ClientCity'] = row[7]
            clients[client_id_2]['ClientState'] = row[8]
            clients[client_id_2]['ClientZip'] = row[9]

    except KeyError:
        pass


# Check other Tabs in original workbook for additional client data
# Match Client ID with Client in the hash table and update missing fields.
client_phone = file['Phone']
for row in client_phone.iter_rows(min_row=2, values_only=True, min_col=1):
    try:
        client_id_2 = row[0]
        if clients[client_id_2]:
            phone_num = str(row[3]) + str(row[4])
            clients[client_id_2]['PhoneNumberWork'] = phone_num

    except KeyError:
        pass


# Append our proper header to the new worksheet
sheet.append(template_header)
sheet.title = 'Clients'

# Add each client to the new spreadsheet
for client in clients:
    client_list = [v for k, v in clients[client].items()]

    # Add client to the worksheet
    sheet.append(client_list)

# Save the worksheet when all is complete
outputFileName = "AccessLiving_modified.xlsx"
workbook.save(filename=outputFileName)
