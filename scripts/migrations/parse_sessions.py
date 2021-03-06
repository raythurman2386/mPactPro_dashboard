from openpyxl import load_workbook, Workbook
from correct_client import fix_data
import sys
import os

# This is the file that will be corrected
# masterFile = sys.argv[1]

# This is the file that will be getting corrected
# file = load_workbook(filename=masterFile, data_only=True)

# case_sheet = file.active

# This will be a new workbook that we save our corrected values to
workbook = load_workbook(filename='HOI_modified.xlsx', data_only=True)
sheet = workbook.create_sheet('Session')

# Create a hashtable to keep track of cases
sessions = {}

template_header = ['SessionDuration', 'CounselorName', 'SessionType', 'ClientNotes', '9series', '10a', '10b', '10c',
                   '10d', '10e', '10f', '10g', '10h', '10i', '10j', '10k', '10l', '10m', 'SessionFee', 'BillableTo',
                   'IfOtherWho']

# Fill our clients hash table with every client
# for row in case_sheet.iter_rows(min_row=2, values_only=True, min_col=2):
#     session_id = None
#     # TODO: IF THE HEADER COLUMNS ARE IN DIFFERENT LOCATIONS, UPDATE THE VALUES FOR THE CLIENT!!!!!
#     session = {
#         'session_duration': None,
#         'counselor_name': None,
#         'client_notes': None,
#         '9_series': None,
#         '10a': None,
#         '10b': None,
#         '10c': None,
#         '10d': None,
#         '10e': None,
#         '10f': None,
#         '10g': None,
#         '10h': None,
#         '10i': None,
#         '10j': None,
#         '10k': None,
#         '10l': None,
#         '10m': None,
#         'session_fee': None,
#         'billable_to': None,
#         'if_other_who': None,
#     }
#     # Save the client by the ID for easy Access
#     sessions[session_id] = session

# Append our proper header to the new worksheet
sheet.append(template_header)
sheet.title = 'Session'

# Add each client to the new spreadsheet
# for session in sessions:
# session_list = [v for k, v in sessions[session].items()]

# TODO: Fix client data to match our requirements
# corrected_session = fix_data(session_list)

# Add corrected client to the worksheet
# sheet.append(corrected_session)

# Save the worksheet when all is complete
workbook.save(filename='HOI_modified_cases.xlsx')
